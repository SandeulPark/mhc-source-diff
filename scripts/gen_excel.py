#!/usr/bin/env python3
"""prd/ 와 svn/ 비교 엑셀 생성

사용법: python scripts/gen_excel.py <target_dir>
예시:   python scripts/gen_excel.py mhcweb
"""

import os
import sys
import json
import datetime
import hashlib
import re
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)


def file_md5(filepath):
    """파일 MD5 체크섬 계산"""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def load_mtime_json(json_path):
    """prd_mtime.json 로드. 없으면 None 반환."""
    if not os.path.isfile(json_path):
        return None
    with open(json_path, "r", encoding="utf-8") as fp:
        data = json.load(fp)
    return data.get("files", {})


MAJOR_VERSION_MAP = {
    45: "1.1", 46: "1.2", 47: "1.3", 48: "1.4", 49: "1.5",
    50: "1.6", 51: "1.7", 52: "1.8", 53: "9", 54: "10",
    55: "11", 56: "12", 57: "13", 58: "14", 59: "15",
    60: "16", 61: "17", 62: "18", 63: "19", 64: "20", 65: "21",
}


def get_class_jdk_version(filepath):
    """class 파일의 major version으로 JDK 버전 반환"""
    try:
        with open(filepath, "rb") as f:
            magic = f.read(4)
            if magic != b'\xca\xfe\xba\xbe':
                return ""
            minor = int.from_bytes(f.read(2), "big")
            major = int.from_bytes(f.read(2), "big")
        return MAJOR_VERSION_MAP.get(major, str(major))
    except (OSError, ValueError):
        return ""


def scan_class_files(root_dir, mtime_data=None):
    """디렉토리 스캔하여 .class 파일 정보 수집.
    mtime_data가 주어지면 해당 딕셔너리에서 mtime을 읽는다."""
    result = {}
    for dirpath, _, filenames in os.walk(root_dir):
        for f in filenames:
            if f.endswith(".class"):
                full = os.path.join(dirpath, f)
                rel = os.path.relpath(full, root_dir)
                rel_posix = rel.replace("\\", "/")
                stat = os.stat(full)
                if mtime_data and rel_posix in mtime_data:
                    mtime = datetime.datetime.strptime(
                        mtime_data[rel_posix], "%Y-%m-%d %H:%M:%S"
                    )
                else:
                    mtime = datetime.datetime.fromtimestamp(stat.st_mtime)
                size = stat.st_size
                md5 = file_md5(full)
                jdk = get_class_jdk_version(full)
                result[rel] = {"mtime": mtime, "size": size, "md5": md5, "jdk": jdk}
    return result


def javap_signature(filepath):
    """javap 출력에서 컴파일러 버전 차이를 무시한 시그니처 해시 반환"""
    try:
        result = subprocess.run(
            ["javap", "-c", "-p", filepath],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode != 0:
            return None
        lines = []
        for line in result.stdout.splitlines():
            if any(skip in line for skip in [
                "Classfile", "Compiled from"
            ]):
                continue
            # constant pool 인덱스 번호 제거 (#숫자 // → //)
            line = re.sub(r'#\d+(?:,\s*\d+)?\s*//', '//', line)
            lines.append(line)
        content = "\n".join(lines)
        return hashlib.md5(content.encode("utf-8")).hexdigest()
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return None


def javap_compare(rel_path, classes_dir, svn_dir):
    """MD5가 다른 파일에 대해 javap 비교 수행. 반환: '코드변경' 또는 '변경없음'"""
    prd_path = os.path.join(classes_dir, rel_path)
    svn_path = os.path.join(svn_dir, rel_path)
    prd_sig = javap_signature(prd_path)
    svn_sig = javap_signature(svn_path)
    if prd_sig is None or svn_sig is None:
        return "javap실패"
    if prd_sig == svn_sig:
        return "변경없음"
    else:
        return "코드변경"


def split_path(rel_path):
    """경로를 depth별로 분리 (kr/go/mhc 이후 기준)"""
    parts = rel_path.replace("\\", "/").split("/")
    # kr/go/mhc/ 이후부터 의미 있는 depth
    # parts: kr, go, mhc, <영역>, <모듈>, <계층...>, 파일명
    if len(parts) < 4:
        return ("", "", "", "", rel_path.split("/")[-1])

    after_mhc = parts[3:]  # mhc 이후
    filename = after_mhc[-1]
    path_parts = after_mhc[:-1]

    area = path_parts[0] if len(path_parts) > 0 else ""
    module = path_parts[1] if len(path_parts) > 1 else ""
    layer = "/".join(path_parts[2:]) if len(path_parts) > 2 else ""

    return (area, module, layer, filename)


def compare_status(rel_path, classes_data, svn_data):
    """MD5 체크섬 기반 비교 상태 반환"""
    in_classes = rel_path in classes_data
    in_svn = rel_path in svn_data

    if in_classes and in_svn:
        if classes_data[rel_path]["md5"] == svn_data[rel_path]["md5"]:
            return "변경없음"
        else:
            return "변경됨"
    elif in_classes and not in_svn:
        return "prd에만 존재"
    else:
        return "SVN에만 존재"


def main():
    if len(sys.argv) < 2:
        print("사용법: python scripts/gen_excel.py <target_dir>")
        print("예시:   python scripts/gen_excel.py mhcweb")
        sys.exit(1)

    target = sys.argv[1]
    classes_dir = os.path.join(PROJECT_ROOT, target, "prd")
    svn_dir = os.path.join(PROJECT_ROOT, target, "svn")
    prd_mtime_json = os.path.join(PROJECT_ROOT, target, "prd_mtime.json")
    svn_mtime_json = os.path.join(PROJECT_ROOT, target, "svn_mtime.json")

    if not os.path.isdir(classes_dir):
        print(f"오류: {classes_dir} 디렉토리가 존재하지 않습니다.")
        sys.exit(1)
    if not os.path.isdir(svn_dir):
        print(f"오류: {svn_dir} 디렉토리가 존재하지 않습니다.")
        sys.exit(1)

    print("파일 스캔 중...")
    prd_mtime_data = load_mtime_json(prd_mtime_json)
    if prd_mtime_data:
        print(f"prd_mtime.json 로드 완료: {len(prd_mtime_data)}개 엔트리")
    else:
        print("prd_mtime.json 없음 → os.stat() 사용")
    svn_mtime_data = load_mtime_json(svn_mtime_json)
    if svn_mtime_data:
        print(f"svn_mtime.json 로드 완료: {len(svn_mtime_data)}개 엔트리")
    else:
        print("svn_mtime.json 없음 → os.stat() 사용")
    classes_data = scan_class_files(classes_dir, mtime_data=prd_mtime_data)
    svn_data = scan_class_files(svn_dir, mtime_data=svn_mtime_data)

    # 전체 경로 합집합
    all_paths = sorted(set(list(classes_data.keys()) + list(svn_data.keys())))

    print(
        f"prd: {len(classes_data)}개, svn: {len(svn_data)}개, 합계(유니크): {len(all_paths)}개"
    )

    # 1단계: MD5 비교로 "변경됨" 파일 식별
    changed_paths = [
        p for p in all_paths
        if p in classes_data and p in svn_data
        and classes_data[p]["md5"] != svn_data[p]["md5"]
    ]

    # 2단계: javap 하이브리드 비교
    javap_results = {}
    if changed_paths:
        print(f"\njavap 비교 시작 ({len(changed_paths)}개 파일)...")
        for i, rel_path in enumerate(changed_paths, 1):
            print(f"  [{i}/{len(changed_paths)}] {rel_path}", end="")
            result = javap_compare(rel_path, classes_dir, svn_dir)
            javap_results[rel_path] = result
            print(f" → {result}")
        meta_only = sum(1 for v in javap_results.values() if v == "변경없음")
        code_chg = sum(1 for v in javap_results.values() if v == "코드변경")
        javap_fail = sum(1 for v in javap_results.values() if v == "javap실패")
        print(f"javap 비교 완료: 코드변경 {code_chg}개, 변경없음 {meta_only}개, javap실패 {javap_fail}개\n")

    wb = Workbook()
    ws = wb.active
    ws.title = "prd vs svn 비교"

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 상태별 색상
    status_fills = {
        "변경없음": PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        ),
        "코드변경": PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        ),
        "javap실패": PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        ),
        "prd에만 존재": PatternFill(
            start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"
        ),
        "SVN에만 존재": PatternFill(
            start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
        ),
    }

    # 헤더
    headers = [
        "No",
        "영역",  # common / mhcapp / mhcweb
        "모듈",  # cm, gn, mr, sv, tg, ...
        "계층",  # controller / service / service/impl / util / crontab
        "파일명",
        "전체경로",
        "비교상태",
        "최신 버전",
        "prd JDK",
        "prd 수정일시",
        "prd 파일크기(bytes)",
        "prd MD5",
        "svn JDK",
        "svn 수정일시",
        "svn 파일크기(bytes)",
        "svn MD5",
        "크기차이(bytes)",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 데이터 작성
    for idx, rel_path in enumerate(all_paths, 1):
        row = idx + 1
        area, module, layer, filename = split_path(rel_path)
        status = compare_status(rel_path, classes_data, svn_data)

        # javap 결과로 "변경됨"을 세분화
        if status == "변경됨" and rel_path in javap_results:
            status = javap_results[rel_path]

        cls_info = classes_data.get(rel_path)
        svn_info = svn_data.get(rel_path)

        cls_jdk = cls_info["jdk"] if cls_info else ""
        cls_mtime = cls_info["mtime"].strftime("%Y-%m-%d %H:%M:%S") if cls_info else ""
        cls_size = cls_info["size"] if cls_info else ""
        cls_md5 = cls_info["md5"] if cls_info else ""
        svn_jdk = svn_info["jdk"] if svn_info else ""
        svn_mtime = svn_info["mtime"].strftime("%Y-%m-%d %H:%M:%S") if svn_info else ""
        svn_size = svn_info["size"] if svn_info else ""
        svn_md5 = svn_info["md5"] if svn_info else ""

        size_diff = ""
        if cls_info and svn_info:
            size_diff = cls_info["size"] - svn_info["size"]

        # 최신 버전 판별
        if status in ("코드변경", "변경없음", "javap실패") and cls_info and svn_info:
            if cls_info["mtime"] > svn_info["mtime"]:
                newer = "prd"
            elif cls_info["mtime"] < svn_info["mtime"]:
                newer = "svn"
            else:
                newer = "동일"
        else:
            newer = ""

        values = [
            idx,
            area,
            module,
            layer,
            filename,
            rel_path.replace("\\", "/"),
            status,
            newer,
            cls_jdk,
            cls_mtime,
            cls_size,
            cls_md5,
            svn_jdk,
            svn_mtime,
            svn_size,
            svn_md5,
            size_diff,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col == 7:  # 비교상태 컬럼
                cell.fill = status_fills.get(status, PatternFill())
                cell.alignment = Alignment(horizontal="center")
            if col == 8:  # 최신 버전 컬럼
                cell.alignment = Alignment(horizontal="center")
            if col == 1:  # No
                cell.alignment = Alignment(horizontal="center")

    # 열 너비 조정
    col_widths = [6, 12, 12, 18, 45, 70, 16, 12, 10, 22, 20, 36, 10, 22, 24, 36, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 필터 설정
    ws.auto_filter.ref = f"A1:Q{len(all_paths) + 1}"

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    ws2.cell(row=1, column=1, value="비교 상태").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="파일 수").font = Font(bold=True)

    status_count = {}
    for rel_path in all_paths:
        s = compare_status(rel_path, classes_data, svn_data)
        if s == "변경됨" and rel_path in javap_results:
            s = javap_results[rel_path]
        status_count[s] = status_count.get(s, 0) + 1

    for i, (s, cnt) in enumerate(sorted(status_count.items()), 2):
        ws2.cell(row=i, column=1, value=s)
        ws2.cell(row=i, column=2, value=cnt)
        ws2.cell(row=i, column=1).fill = status_fills.get(s, PatternFill())

    ws2.cell(row=len(status_count) + 2, column=1, value="합계").font = Font(bold=True)
    ws2.cell(row=len(status_count) + 2, column=2, value=len(all_paths)).font = Font(
        bold=True
    )

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12

    # 저장
    output = os.path.join(PROJECT_ROOT, target, "classes_comparison.xlsx")
    wb.save(output)
    print(f"엑셀 파일 생성 완료: {output}")


if __name__ == "__main__":
    main()

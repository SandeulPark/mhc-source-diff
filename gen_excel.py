#!/usr/bin/env python3
"""classes/ 와 svn/ 비교 엑셀 생성"""

import os
import datetime
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CLASSES_DIR = os.path.join(BASE_DIR, "prd")
SVN_DIR = os.path.join(BASE_DIR, "svn")

def file_md5(filepath):
    """파일 MD5 체크섬 계산"""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def scan_class_files(root_dir):
    """디렉토리 스캔하여 .class 파일 정보 수집"""
    result = {}
    for dirpath, _, filenames in os.walk(root_dir):
        for f in filenames:
            if f.endswith(".class"):
                full = os.path.join(dirpath, f)
                rel = os.path.relpath(full, root_dir)
                stat = os.stat(full)
                mtime = datetime.datetime.fromtimestamp(stat.st_mtime)
                size = stat.st_size
                md5 = file_md5(full)
                result[rel] = {"mtime": mtime, "size": size, "md5": md5}
    return result

def split_path(rel_path):
    """경로를 depth별로 분리 (kr/go/mhc 이후 기준)"""
    parts = rel_path.replace("\\", "/").split("/")
    # kr/go/mhc/ 이후부터 의미 있는 depth
    # parts: kr, go, mhc, <영역>, <모듈>, <계층...>, 파일명
    if len(parts) < 4:
        return ("", "", "", "", rel_path.split("/")[-1])

    after_mhc = parts[3:]  # mhc 이후
    filename = after_mhc[-1]
    path_parts = after_mhc[:-1]

    area = path_parts[0] if len(path_parts) > 0 else ""
    module = path_parts[1] if len(path_parts) > 1 else ""
    layer = "/".join(path_parts[2:]) if len(path_parts) > 2 else ""

    return (area, module, layer, filename)

def compare_status(rel_path, classes_data, svn_data):
    """MD5 체크섬 기반 비교 상태 반환"""
    in_classes = rel_path in classes_data
    in_svn = rel_path in svn_data

    if in_classes and in_svn:
        if classes_data[rel_path]["md5"] == svn_data[rel_path]["md5"]:
            return "동일"
        else:
            return "변경됨"
    elif in_classes and not in_svn:
        return "prd에만 존재"
    else:
        return "SVN에만 존재"

def main():
    print("파일 스캔 중...")
    classes_data = scan_class_files(CLASSES_DIR)
    svn_data = scan_class_files(SVN_DIR)

    # 전체 경로 합집합
    all_paths = sorted(set(list(classes_data.keys()) + list(svn_data.keys())))

    print(f"prd: {len(classes_data)}개, svn: {len(svn_data)}개, 합계(유니크): {len(all_paths)}개")

    wb = Workbook()
    ws = wb.active
    ws.title = "prd vs svn 비교"

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 상태별 색상
    status_fills = {
        "동일": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "변경됨": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "prd에만 존재": PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
        "SVN에만 존재": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    }

    # 헤더
    headers = [
        "No",
        "영역",        # common / mhcapp / mhcweb
        "모듈",        # cm, gn, mr, sv, tg, ...
        "계층",        # controller / service / service/impl / util / crontab
        "파일명",
        "전체경로",
        "비교상태",
        "prd 수정일시",
        "prd 파일크기(bytes)",
        "prd MD5",
        "svn 수정일시",
        "svn 파일크기(bytes)",
        "svn MD5",
        "크기차이(bytes)",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 데이터 작성
    for idx, rel_path in enumerate(all_paths, 1):
        row = idx + 1
        area, module, layer, filename = split_path(rel_path)
        status = compare_status(rel_path, classes_data, svn_data)

        cls_info = classes_data.get(rel_path)
        svn_info = svn_data.get(rel_path)

        cls_mtime = cls_info["mtime"].strftime("%Y-%m-%d %H:%M:%S") if cls_info else ""
        cls_size = cls_info["size"] if cls_info else ""
        cls_md5 = cls_info["md5"] if cls_info else ""
        svn_mtime = svn_info["mtime"].strftime("%Y-%m-%d %H:%M:%S") if svn_info else ""
        svn_size = svn_info["size"] if svn_info else ""
        svn_md5 = svn_info["md5"] if svn_info else ""

        size_diff = ""
        if cls_info and svn_info:
            size_diff = cls_info["size"] - svn_info["size"]

        values = [
            idx,
            area,
            module,
            layer,
            filename,
            rel_path.replace("\\", "/"),
            status,
            cls_mtime,
            cls_size,
            cls_md5,
            svn_mtime,
            svn_size,
            svn_md5,
            size_diff,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col == 7:  # 비교상태 컬럼
                cell.fill = status_fills.get(status, PatternFill())
                cell.alignment = Alignment(horizontal="center")
            if col == 1:  # No
                cell.alignment = Alignment(horizontal="center")

    # 열 너비 조정
    col_widths = [6, 12, 12, 18, 45, 70, 16, 22, 20, 36, 22, 24, 36, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 필터 설정
    ws.auto_filter.ref = f"A1:N{len(all_paths) + 1}"

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    ws2.cell(row=1, column=1, value="비교 상태").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="파일 수").font = Font(bold=True)

    status_count = {}
    for rel_path in all_paths:
        s = compare_status(rel_path, classes_data, svn_data)
        status_count[s] = status_count.get(s, 0) + 1

    for i, (s, cnt) in enumerate(sorted(status_count.items()), 2):
        ws2.cell(row=i, column=1, value=s)
        ws2.cell(row=i, column=2, value=cnt)
        ws2.cell(row=i, column=1).fill = status_fills.get(s, PatternFill())

    ws2.cell(row=len(status_count) + 2, column=1, value="합계").font = Font(bold=True)
    ws2.cell(row=len(status_count) + 2, column=2, value=len(all_paths)).font = Font(bold=True)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12

    # 저장
    output = os.path.join(BASE_DIR, "classes_comparison.xlsx")
    wb.save(output)
    print(f"엑셀 파일 생성 완료: {output}")

if __name__ == "__main__":
    main()
